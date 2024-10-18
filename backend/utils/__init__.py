import os
import csv
from flask import current_app as app
from pandas import DataFrame
from qrcode import QRCode, constants
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from fpdf import FPDF
from datetime import datetime, timedelta


def create_file_path(meeting_id, file_name):
    return os.path.join(app.root_path, f"static/meeting_{meeting_id}_{file_name}")


def create_qr_image(data):
    qr = QRCode(
        version=1,
        error_correction=constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill="black", back_color="white")
    return img


def generate_qr_code(url, meeting_id):
    img = create_qr_image(str(url) + str(meeting_id))
    qr_file = create_file_path(meeting_id, "qr.png")
    # Save image to the file
    img.save(qr_file)
    return qr_file


def convert_data_to_excel(file, data):
    df = DataFrame(data)
    df.to_excel(file, index=False)
    # Read excel file using openpyxl
    workbook = load_workbook(file)
    worksheet = workbook.active

    # Adjust column width
    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        # Add extra padding to the width
        worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

    # Bold title row
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    # Align cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
    return workbook


def generate_excel_file(data, meeting_id):
    file = create_file_path(meeting_id, "report.xlsx")
    workbook = convert_data_to_excel(file, data)
    workbook.save(file)
    return file


def add_logo(pdf):
    print("Width:", pdf.w)
    pdf.image(os.path.join(app.root_path, "static/logo.png"), x=122, y=8, w=50)
    pdf.ln(30)


def add_pdf_title(pdf, meeting):
    pdf.set_font("Times", "B", 20)
    pdf.cell(
        w=None,
        h=None,
        text="Meeting Report",
        center=True,
    )
    pdf.ln(10)
    pdf.set_font("Times", "B", 14)

    pdf.cell(0, 10, txt=f"Title: {meeting['title']}", ln=True)
    pdf.cell(0, 10, txt=f"Location: {meeting['location']}", ln=True)
    pdf.cell(0, 10, txt=f"Description: {meeting['description']}", ln=True)
    pdf.cell(0, 10, txt=f"Date: {meeting['meeting_date']}", ln=True)
    pdf.cell(0, 10, txt=f"Start Time: {meeting['start_time']}", ln=True)
    pdf.cell(0, 10, txt=f"End Time: {meeting['end_time']}", ln=True)
    pdf.ln(10)


def add_pdf_body(pdf, attendees):
    convert_json_to_csv(attendees)
    file = os.path.join(app.root_path, "static/data.txt")

    with open(file, encoding="utf8") as csv_file:
        datas = list(csv.reader(csv_file, delimiter=","))

    pdf.set_font("Times", size=12)
    with pdf.table(
        col_widths=(10, 30, 30, 30, 30, 50, 30),
    ) as table:
        for data_row in datas:
            row = table.row()
            for datum in data_row:
                row.cell(datum)


def create_pdf(pdf, meeting, attendees):
    add_logo(pdf)
    add_pdf_title(pdf, meeting)
    add_pdf_body(pdf, attendees)
    # remove CSV file
    os.remove(os.path.join(app.root_path, "static/data.txt"))


def generate_pdf_file(meeting, attendees):
    file = create_file_path(meeting.get("id"), "report.pdf")
    pdf = FPDF(orientation="L", format="A4")
    pdf.add_page()
    create_pdf(pdf, meeting, attendees)
    pdf.output(file)
    return file


def convert_json_to_csv(data):
    output_csv_file = os.path.join(app.root_path, "static/data.txt")
    with open(output_csv_file, "w", newline="") as csv_file:
        headers = {
            "id": "ID",
            "first_name": "First Name",
            "last_name": "Last Name",
            "organization": "Organization",
            "designation": "Designation",
            "email": "Email",
            "phone": "Phone",
        }

        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writerow(headers)
        for item in data:
            unwanted_columns = ("meeting_id", "created_on", "updated_on")
            for key in unwanted_columns:
                item.pop(key, None)
            writer.writerow(item)


def parse_date(date_string):
    if isinstance(date_string, str):
        return datetime.strptime(date_string, "%Y-%m-%d").date()


def parse_time(time_string):
    if isinstance(time_string, str):
        return datetime.strptime(time_string, "%H:%M:%S").time()


def check_time_difference(time_string):
    if isinstance(time_string, timedelta):
        return (datetime.min + time_string).time()


def combine_date_time(date_string, time_string):
    return datetime.combine(date_string, parse_time(time_string))


statements = {
    "organizations": "CREATE TABLE IF NOT EXISTS organizations (id INT AUTO_INCREMENT PRIMARY KEY,name VARCHAR(100) NOT NULL,description TEXT,created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "resources": "CREATE TABLE IF NOT EXISTS resources (id INT AUTO_INCREMENT PRIMARY KEY,name VARCHAR(100) NOT NULL, description VARCHAR(100),quantity INT NOT NULL,status ENUM('available', 'unavailable') DEFAULT 'available', created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "users": "CREATE TABLE IF NOT EXISTS users (id INT AUTO_INCREMENT PRIMARY KEY,first_name VARCHAR(100) NOT NULL,last_name VARCHAR(100) NOT NULL, organization VARCHAR(100) NOT NULL, designation VARCHAR(100) NOT NULL, email VARCHAR(100) NOT NULL, phone VARCHAR(30), password VARCHAR(255) NOT NULL, created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "boardrooms": "CREATE TABLE IF NOT EXISTS boardrooms (id INT AUTO_INCREMENT PRIMARY KEY,name VARCHAR(100) NOT NULL,capacity INT NOT NULL,location VARCHAR(100) NOT NULL,description VARCHAR(100),status ENUM('available','unavailable') DEFAULT 'available', created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "meetings": "CREATE TABLE IF NOT EXISTS meetings (id INT AUTO_INCREMENT PRIMARY KEY,title VARCHAR(100) NOT NULL,description TEXT,meeting_date DATE NOT NULL,start_time TIME NOT NULL,end_time TIME NOT NULL,boardroom_id INT,organization_id INT NOT NULL, resources_id JSON,status ENUM('draft', 'ongoing', 'complete', 'rescheduled','pending', 'cancelled') DEFAULT 'pending',location TEXT, FOREIGN KEY (boardroom_id) REFERENCES boardrooms(id),created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "attendees": "CREATE TABLE IF NOT EXISTS attendees (id INT AUTO_INCREMENT PRIMARY KEY,first_name VARCHAR(100) NOT NULL,last_name VARCHAR(100) NOT NULL, organization VARCHAR(100), designation VARCHAR(100) NOT NULL, email VARCHAR(100) NOT NULL, phone VARCHAR(30), meeting_id INT NOT NULL,FOREIGN KEY (meeting_id) REFERENCES meetings(id),created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP)",
    "roles": "CREATE TABLE IF NOT EXISTS roles (id INT AUTO_INCREMENT PRIMARY KEY,name VARCHAR(50) NOT NULL UNIQUE,description TEXT,created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "users_roles": "CREATE TABLE IF NOT EXISTS users_roles (user_id INT NOT NULL,role_id INT NOT NULL,PRIMARY KEY (user_id, role_id),FOREIGN KEY (user_id) REFERENCES users(id),FOREIGN KEY (role_id) REFERENCES roles(id),created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP );",
    "permissions": "CREATE TABLE IF NOT EXISTS permissions (id INT AUTO_INCREMENT PRIMARY KEY,name VARCHAR(50) NOT NULL UNIQUE,created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "roles_permissions": "CREATE TABLE IF NOT EXISTS roles_permissions (role_id INT NOT NULL,permission_id INT NOT NULL,PRIMARY KEY (role_id, permission_id),FOREIGN KEY (role_id) REFERENCES roles(id),FOREIGN KEY (permission_id) REFERENCES permissions(id),created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "locations": "CREATE TABLE IF NOT EXISTS locations (id INT AUTO_INCREMENT PRIMARY KEY,county VARCHAR(254) NOT NULL,town VARCHAR(254) NOT NULL,created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,updated_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP);",
    "alter_meetings": "SET @tables = (SELECT COUNT(*) FROM information_schema.tables WHERE table_name = 'meetings'); SET @columns = (SELECT COUNT(*) FROM information_schema.columns WHERE table_name = 'meetings' AND column_name IN ('longitude', 'latitude', 'county', 'town')); IF @tables > 0 AND @columns = 0 THEN ALTER TABLE meetings ADD COLUMN longitude DECIMAL(10, 8) AFTER location, ADD COLUMN latitude DECIMAL(10, 8) AFTER longitude, ADD COLUMN county VARCHAR(254) AFTER latitude, ADD COLUMN town VARCHAR(254) AFTER county; END IF;",

}
