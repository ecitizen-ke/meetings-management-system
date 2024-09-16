# Utility functions for QR code and report generation
import qrcode
import os
import MySQLdb
import pandas as pd
from flask import url_for, current_app as app
from .models import get_db_connection
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from fpdf import FPDF

def generate_qr_code(meeting_id):
    url = "http://localhost:5173/meeting/" + str(meeting_id)  # For testing purposes
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)

    img = qr.make_image(fill='black', back_color='white')
    qr_path = os.path.join(app.root_path, f'static/meeting_{meeting_id}_qr.png')
    img.save(qr_path)
    return qr_path

def generate_excel_report(meeting_id):
    connection = get_db_connection()
    cursor = connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT first_name, last_name, email, phone, department, designation FROM attendees WHERE meeting_id = %s", (meeting_id,))
    attendees = cursor.fetchall()
    cursor.close()
    connection.close()

    # Convert to DataFrame
    df = pd.DataFrame(attendees)
    excel_path = os.path.join(app.root_path, f'static/meeting_{meeting_id}_report.xlsx')

    # Write DataFrame to Excel file without index
    df.to_excel(excel_path, index=False)

    # Open the Excel file using openpyxl to format it
    wb = load_workbook(excel_path)
    ws = wb.active

    # Adjust column widths based on the max length of data in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Try to get the length of the cell's value
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra padding to the width
        ws.column_dimensions[column].width = adjusted_width

    # Optionally, set a title row style (bold)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Define border style (thin border)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Center align all cells and apply borders
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Save the formatted Excel file
    wb.save(excel_path)

    return excel_path

def generate_pdf_report(meeting_id):
    connection = get_db_connection()
    cursor = connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT title, description, meeting_date, start_time, end_time, boardroom_id FROM meetings WHERE id = %s", (meeting_id,))
    meeting = cursor.fetchone()
    cursor.execute("SELECT name FROM boardrooms WHERE id = %s", (meeting['boardroom_id'],))
    boardroom = cursor.fetchone()
    cursor.execute("SELECT first_name, last_name, email, phone, department, designation FROM attendees WHERE meeting_id = %s", (meeting_id,))
    attendees = cursor.fetchall()
    cursor.close()
    connection.close()

    pdf_path = os.path.join(app.root_path, f'static/meeting_{meeting_id}_report.pdf')
    pdf = FPDF()
    pdf.add_page()

    # Add organization logo
    pdf.image(os.path.join(app.root_path, 'static/logo.png'), x=10, y=8, w=33)

    # Meeting details
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Meeting Report", ln=True, align='C')
    pdf.ln(10)

    pdf.set_font("Arial", size=10)
    pdf.cell(0, 10, txt=f"Title: {meeting['title']}", ln=True)
    pdf.cell(0, 10, txt=f"Location: {boardroom['name']}", ln=True)
    pdf.cell(0, 10, txt=f"Description: {meeting['description']}", ln=True)
    pdf.cell(0, 10, txt=f"Date: {meeting['meeting_date']}", ln=True)
    pdf.cell(0, 10, txt=f"Start Time: {meeting['start_time']}", ln=True)
    pdf.cell(0, 10, txt=f"End Time: {meeting['end_time']}", ln=True)
    pdf.ln(10)

    # Attendees
    pdf.set_font("Arial", size=10)
    pdf.cell(200, 10, txt="Attendees:", ln=True)
    pdf.ln(5)

    pdf.set_font("Arial", size=9)
    for attendee in attendees:
        pdf.cell(0, 10, txt=f"{attendee['first_name']} {attendee['last_name']}", ln=True)
        pdf.cell(0, 10, txt=f"Email: {attendee['email']}", ln=True)
        pdf.cell(0, 10, txt=f"Phone: {attendee['phone']}", ln=True)
        pdf.cell(0, 10, txt=f"Department: {attendee['department']}", ln=True)
        pdf.cell(0, 10, txt=f"Designation: {attendee['designation']}", ln=True)
        pdf.ln(5)

    pdf.output(pdf_path)
    return pdf_path
